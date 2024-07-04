from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'

USERS_FILE = 'employees.xlsx'
AVAILABILITY_FILE = 'availability.xlsx'

def init_excel_files():
    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Full Name', 'Employment Type', 'Encrypted Password'])
        wb.save(USERS_FILE)

    if not os.path.exists(AVAILABILITY_FILE):
        wb = Workbook()
        ws = wb.active
        headers = ['Full Name', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        ws.append(headers)
        format_excel_sheet(ws)
        wb.save(AVAILABILITY_FILE)

def format_excel_sheet(ws):
    # Format header row
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

def update_excel_formatting():
    wb = load_workbook(AVAILABILITY_FILE)
    ws = wb.active
    format_excel_sheet(ws)
    wb.save(AVAILABILITY_FILE)

init_excel_files()
update_excel_formatting()

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        firstname = request.form['firstname']
        lastname = request.form['lastname']
        full_name = f"{firstname} {lastname}"
        employment_type = request.form['employment_type']
        password = request.form['password']
        
        hashed_password = generate_password_hash(password)
        
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        ws.append([full_name, employment_type, hashed_password])
        wb.save(USERS_FILE)
        
        flash('Account created successfully. Please log in.', 'success')
        return redirect(url_for('login'))
    
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        firstname = request.form['firstname']
        lastname = request.form['lastname']
        full_name = f"{firstname} {lastname}"
        password = request.form['password']
        
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == full_name and check_password_hash(row[2], password):
                session['user_full_name'] = full_name
                flash('Logged in successfully.', 'success')
                return redirect(url_for('availability'))
        
        flash('Invalid credentials. Please try again.', 'error')
    
    return render_template('login.html')

@app.route('/availability', methods=['GET', 'POST'])
def availability():
    if 'user_full_name' not in session:
        return redirect(url_for('login'))
    
    user_full_name = session['user_full_name']
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    if request.method == 'POST':
        try:
            wb = load_workbook(AVAILABILITY_FILE)
            ws = wb.active
            
            user_row = None
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_full_name:
                    user_row = row
                    break
            
            if user_row is None:
                ws.append([user_full_name] + [''] * 7)
                user_row = ws[ws.max_row]
            
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            
            for i, day in enumerate(days):
                if day not in request.form:
                    user_row[i+1].value = 'x'
                    user_row[i+1].fill = red_fill
                else:
                    user_row[i+1].value = ''
                    user_row[i+1].fill = PatternFill(fill_type=None)
                
                # Center-align the cell content
                user_row[i+1].alignment = Alignment(horizontal='center', vertical='center')
            
            format_excel_sheet(ws)
            wb.save(AVAILABILITY_FILE)
            return jsonify({'success': True, 'message': 'Availability updated successfully!'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    
    # Get current availability
    wb = load_workbook(AVAILABILITY_FILE)
    ws = wb.active
    availabilities = {day: '' for day in days}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_full_name:
            availabilities = dict(zip(days, row[1:]))
            break
    
    return render_template('availability.html', user_full_name=user_full_name, days=days, availabilities=availabilities)

@app.route('/logout')
def logout():
    session.pop('user_full_name', None)
    flash('Logged out successfully.', 'success')
    return redirect(url_for('home'))

@app.route('/export')
def export():
    if not os.path.exists(AVAILABILITY_FILE):
        flash('No availability data to export.', 'error')
        return redirect(url_for('availability'))
    
    return send_file(AVAILABILITY_FILE, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
